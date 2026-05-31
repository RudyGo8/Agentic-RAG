# -*- coding: utf-8 -*-
"""Offline RAGAS evaluation for retrieval and generation quality."""

from __future__ import annotations

import asyncio
import csv
import json
import math
import sys
import types
from datetime import datetime
from pathlib import Path

import openpyxl
from openai import AsyncOpenAI


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "backend"))

DATASET = ROOT / "evals" / "datasets" / "ragas_testset_十五五计划.xlsx"
REPORT_DIR = ROOT / "evals" / "experiments"
LIMIT = 0
MAX_CONTEXT_CHARS = 1200
METRIC_NAMES = ["context_precision", "context_recall", "faithfulness", "context_relevance"]


def clean(value) -> str:
    return str(value or "").strip()


# 读取 Excel 测试集
def load_cases() -> list[dict[str, str]]:
    workbook = openpyxl.load_workbook(DATASET, read_only=True, data_only=True)
    sheet = workbook["ragas_testset"] if "ragas_testset" in workbook.sheetnames else workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    columns = {clean(name): index for index, name in enumerate(rows[0])}

    cases = []
    for row_num, row in enumerate(rows[1:], start=2):
        question = clean(row[columns["question"]])
        if question:
            cases.append(
                {
                    "question_id": clean(row[columns.get("question_id", -1)]) or f"row_{row_num}",
                    "question": question,
                    "reference": clean(row[columns["answer"]]),
                }
            )
    return cases[:LIMIT] if LIMIT else cases


# 调用项目 RAG 图进行检索
def run_rag(question: str) -> tuple[list[dict], list[str]]:
    from app.rag.graph import run_rag_graph

    state = run_rag_graph(question)
    docs = state.get("docs", []) if isinstance(state, dict) else []
    contexts = [
        clean(doc.get("text") or doc.get("page_content") or doc.get("content"))[:MAX_CONTEXT_CHARS]
        for doc in docs
    ]
    return docs, [context for context in contexts if context]


# 基于检索结果生成回答
def generate_answer(question: str, docs: list[dict]) -> str:
    from app.agent.factory import get_model
    from app.rag.formatter import format_docs

    if not docs:
        return "未检索到可用资料，无法基于文档回答。"

    context = "\n\n".join(format_docs(docs))
    prompt = f"""请只依据下面的检索资料回答问题。
如果资料不足，请明确说明资料不足，不要编造。

检索资料：
{context}

问题：
{question}

回答："""
    return clean(get_model().invoke(prompt).content)


# 构建 RAGAS 评测指标
def build_metrics() -> dict:
    datasets_stub = types.ModuleType("datasets")
    datasets_stub.Dataset = object
    sys.modules.setdefault("datasets", datasets_stub)

    from app.core.config import RAGAS_API_KEY, RAGAS_BASE_URL, RAGAS_LLM_MODEL
    from ragas.llms import llm_factory
    from ragas.metrics.collections import (
        ContextPrecisionWithReference,
        ContextRecall,
        ContextRelevance,
        Faithfulness,
    )

    if not RAGAS_API_KEY:
        raise RuntimeError("Missing RAGAS_API_KEY")

    llm = llm_factory(
        RAGAS_LLM_MODEL,
        client=AsyncOpenAI(api_key=RAGAS_API_KEY, base_url=RAGAS_BASE_URL),
    )
    return {
        "context_precision": ContextPrecisionWithReference(llm=llm),
        "context_recall": ContextRecall(llm=llm),
        "faithfulness": Faithfulness(llm=llm),
        "context_relevance": ContextRelevance(llm=llm),
    }


# 单个指标安全打分，失败时返回 None
async def score(name: str, metric, **kwargs) -> float | None:
    try:
        return float((await metric.ascore(**kwargs)).value)
    except Exception as exc:
        print(f"[warn] {name} failed: {exc}")
        return None


# 单条样本评测
async def evaluate(case: dict[str, str], metrics: dict) -> dict:
    docs, contexts = run_rag(case["question"])
    response = generate_answer(case["question"], docs)
    row = {
        "question_id": case["question_id"],
        "question": case["question"],
        "reference": case["reference"],
        "response": response,
        "retrieved_count": len(contexts),
    }

    if not contexts:
        return row | {name: None for name in METRIC_NAMES}

    row["context_precision"] = await score(
        "context_precision",
        metrics["context_precision"],
        user_input=case["question"],
        reference=case["reference"],
        retrieved_contexts=contexts,
    )
    row["context_recall"] = await score(
        "context_recall",
        metrics["context_recall"],
        user_input=case["question"],
        reference=case["reference"],
        retrieved_contexts=contexts,
    )
    row["faithfulness"] = await score(
        "faithfulness",
        metrics["faithfulness"],
        user_input=case["question"],
        response=response,
        retrieved_contexts=contexts,
    )
    row["context_relevance"] = await score(
        "context_relevance",
        metrics["context_relevance"],
        user_input=case["question"],
        retrieved_contexts=contexts,
    )
    return row


# 保存 CSV 明细和 summary 汇总
def save(rows: list[dict]) -> None:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    run_name = datetime.now().strftime("ragas_%Y%m%d_%H%M%S")
    csv_path = REPORT_DIR / f"{run_name}.csv"
    summary_path = REPORT_DIR / f"{run_name}_summary.json"

    fields = ["question_id", "question", "reference", "response", "retrieved_count", *METRIC_NAMES]
    with csv_path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)

    summary = {}
    for name in METRIC_NAMES:
        values = [row[name] for row in rows if isinstance(row.get(name), (int, float)) and not math.isnan(row[name])]
        summary[name] = sum(values) / len(values) if values else None

    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print("\n===== RAGAS Summary =====")
    for name, value in summary.items():
        print(f"{name}: {value}")
    print(f"\nCSV: {csv_path}")
    print(f"Summary: {summary_path}")


# 主流程
async def main() -> None:
    metrics = build_metrics()
    rows = []
    cases = load_cases()
    for index, case in enumerate(cases, start=1):
        print(f"[{index}/{len(cases)}] {case['question_id']} {case['question']}")
        rows.append(await evaluate(case, metrics))
    save(rows)


if __name__ == "__main__":
    asyncio.run(main())
